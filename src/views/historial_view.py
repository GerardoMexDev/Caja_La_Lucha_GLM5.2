import customtkinter as ctk
from datetime import datetime, timedelta
from typing import Dict, Any, List
from collections import defaultdict
from tkinter import Toplevel, filedialog
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from tkcalendar import Calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from controllers.caja_controller import CajaController
from controllers.movimiento_controller import MovimientoController
from utils.constants import COLOR_BG, COLOR_CARD, COLOR_TEXT, COLOR_SECONDARY, COLOR_SUCCESS, COLOR_ERROR


class HistorialFrame(ctk.CTkFrame):
    def __init__(self, parent, caja_controller: CajaController, movimiento_controller: MovimientoController) -> None:
        super().__init__(parent, fg_color="transparent")
        self.caja_controller = caja_controller
        self.movimiento_controller = movimiento_controller
        self._ultimos_adelantos: List[Dict[str, Any]] = []
        self._create_widgets()

    def _create_widgets(self) -> None:
        ctk.CTkLabel(self, text="📅 Historial de Cajas", font=ctk.CTkFont(size=24, weight="bold"), text_color=COLOR_TEXT).pack(pady=(10, 20))

        self.tabview = ctk.CTkTabview(self, fg_color=COLOR_CARD)
        self.tabview.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.tab_diario = self.tabview.add("Diario")
        self.tab_semanal = self.tabview.add("Semanal (Lun-Sáb)")
        self.tab_mensual = self.tabview.add("Mensual (Gráfico)")
        self.tab_adelantos = self.tabview.add("Adelantos y Viáticos")

        self._setup_tab_diario()
        self._setup_tab_semanal()
        self._setup_tab_mensual()
        self._setup_tab_adelantos()

    # --- CALENDARIO REUTILIZABLE ---

    def _abrir_calendario(self, entry_widget) -> None:
        top = Toplevel(self.winfo_toplevel())
        top.title("Seleccionar Fecha")
        top.configure(bg="#2b2b2b")
        top.resizable(False, False)
        top.grab_set()
        top.geometry("+%d+%d" % (self.winfo_rootx() + 50, self.winfo_rooty() + 50))

        cal = Calendar(
            top, selectmode="day", date_pattern="yyyy-mm-dd",
            background="#2b2b2b", foreground="white",
            selectbackground="#3498db", selectforeground="white",
            borderwidth=0, headersbackground="#34495e"
        )

        def seleccionar_fecha(event):
            fecha_str = cal.get_date()
            entry_widget.delete(0, "end")
            entry_widget.insert(0, fecha_str)
            top.destroy()

        cal.bind("<<CalendarSelected>>", seleccionar_fecha)
        cal.pack(padx=10, pady=10)

        ctk.CTkButton(top, text="Cancelar", width=100, fg_color="#7f8c8d", command=top.destroy).pack(pady=(0, 10))

    # --- DISEÑO DE PESTAÑAS ---

    def _setup_tab_diario(self) -> None:
        self.tab_diario.grid_columnconfigure(0, weight=1)
        self.tab_diario.grid_rowconfigure(1, weight=1)

        frame_filtros = ctk.CTkFrame(self.tab_diario, fg_color="transparent")
        frame_filtros.grid(row=0, column=0, pady=10, padx=10, sticky="ew")
        frame_filtros.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(frame_filtros, text="Fecha:", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=(0, 10))

        self.fecha_diario = ctk.CTkEntry(frame_filtros, placeholder_text="YYYY-MM-DD", width=150)
        self.fecha_diario.grid(row=0, column=1, sticky="w")
        self.fecha_diario.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.fecha_diario.configure(cursor="hand2")
        self.fecha_diario.bind("<Button-1>", lambda e: self._abrir_calendario(self.fecha_diario))

        ctk.CTkButton(frame_filtros, text="Buscar", width=100, command=self._buscar_diario).grid(row=0, column=2, padx=(10, 0))

        self.container_diario = ctk.CTkScrollableFrame(self.tab_diario, fg_color="transparent")
        self.container_diario.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="nsew")

    def _setup_tab_semanal(self) -> None:
        self.tab_semanal.grid_columnconfigure(0, weight=1)
        self.tab_semanal.grid_rowconfigure(1, weight=1)

        frame_filtros = ctk.CTkFrame(self.tab_semanal, fg_color="transparent")
        frame_filtros.grid(row=0, column=0, pady=10, padx=10, sticky="ew")
        frame_filtros.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(frame_filtros, text="Fecha (para calcular la semana):", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=(0, 10))

        self.fecha_semanal = ctk.CTkEntry(frame_filtros, placeholder_text="YYYY-MM-DD", width=150)
        self.fecha_semanal.grid(row=0, column=1, sticky="w")
        self.fecha_semanal.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.fecha_semanal.configure(cursor="hand2")
        self.fecha_semanal.bind("<Button-1>", lambda e: self._abrir_calendario(self.fecha_semanal))

        ctk.CTkButton(frame_filtros, text="Buscar Semana", width=150, command=self._buscar_semanal).grid(row=0, column=2, padx=(10, 0))

        self.container_semanal = ctk.CTkFrame(self.tab_semanal, fg_color="transparent")
        self.container_semanal.grid(row=1, column=0, padx=20, pady=(0, 10), sticky="nsew")
        self.container_semanal.grid_columnconfigure((0, 1, 2), weight=1)

    def _setup_tab_mensual(self) -> None:
        self.tab_mensual.grid_columnconfigure(0, weight=1)
        self.tab_mensual.grid_rowconfigure(1, weight=1)

        frame_filtros = ctk.CTkFrame(self.tab_mensual, fg_color="transparent")
        frame_filtros.grid(row=0, column=0, pady=10, padx=10, sticky="ew")
        frame_filtros.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(frame_filtros, text="Mes de referencia:", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=(0, 10))

        self.fecha_mensual = ctk.CTkEntry(frame_filtros, placeholder_text="YYYY-MM-DD", width=150)
        self.fecha_mensual.grid(row=0, column=1, sticky="w")
        self.fecha_mensual.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.fecha_mensual.configure(cursor="hand2")
        self.fecha_mensual.bind("<Button-1>", lambda e: self._abrir_calendario(self.fecha_mensual))

        ctk.CTkButton(frame_filtros, text="Ver Gráfico", width=120, command=self._buscar_mensual).grid(row=0, column=2, padx=(10, 0))

        self.container_mensual = ctk.CTkFrame(self.tab_mensual, fg_color="transparent")
        self.container_mensual.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="nsew")

    def _setup_tab_adelantos(self) -> None:
        self.tab_adelantos.grid_columnconfigure(0, weight=1)
        self.tab_adelantos.grid_rowconfigure(2, weight=1)

        frame_filtros = ctk.CTkFrame(self.tab_adelantos, fg_color="transparent")
        frame_filtros.grid(row=0, column=0, pady=10, padx=10, sticky="ew")

        ctk.CTkLabel(frame_filtros, text="Desde:", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=(0, 5))

        self.fecha_desde_av = ctk.CTkEntry(frame_filtros, placeholder_text="YYYY-MM-DD", width=150)
        self.fecha_desde_av.grid(row=0, column=1, padx=(0, 15))
        self.fecha_desde_av.insert(0, (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d"))
        self.fecha_desde_av.configure(cursor="hand2")
        self.fecha_desde_av.bind("<Button-1>", lambda e: self._abrir_calendario(self.fecha_desde_av))

        ctk.CTkLabel(frame_filtros, text="Hasta:", font=ctk.CTkFont(size=14)).grid(row=0, column=2, padx=(0, 5))

        self.fecha_hasta_av = ctk.CTkEntry(frame_filtros, placeholder_text="YYYY-MM-DD", width=150)
        self.fecha_hasta_av.grid(row=0, column=3, padx=(0, 15))
        self.fecha_hasta_av.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.fecha_hasta_av.configure(cursor="hand2")
        self.fecha_hasta_av.bind("<Button-1>", lambda e: self._abrir_calendario(self.fecha_hasta_av))

        ctk.CTkButton(frame_filtros, text="🔍 Buscar", width=120, command=self._buscar_adelantos_viaticos).grid(row=0, column=4)

        frame_excel = ctk.CTkFrame(self.tab_adelantos, fg_color="transparent")
        frame_excel.grid(row=1, column=0, pady=(0, 5), padx=10, sticky="e")

        self.btn_excel_adelantos = ctk.CTkButton(
            frame_excel, text="📥 Exportar Adelantos a Excel", width=250, height=35,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#1e8449", hover_color="#196f3d",
            command=self._exportar_adelantos_excel,
            state="disabled"
        )
        self.btn_excel_adelantos.pack(side="right")

        self.container_av = ctk.CTkScrollableFrame(self.tab_adelantos, fg_color="transparent")
        self.container_av.grid(row=2, column=0, padx=10, pady=(0, 10), sticky="nsew")

    # --- LÓGICA DE BÚSQUEDA Y RENDERIZADO ---

    def _limpiar_frame(self, frame: ctk.CTkFrame) -> None:
        for widget in frame.winfo_children():
            widget.destroy()

    def _buscar_diario(self) -> None:
        fecha = self.fecha_diario.get()
        self._limpiar_frame(self.container_diario)

        cajas = self.caja_controller.obtener_cajas_por_fecha(fecha)

        if not cajas:
            ctk.CTkLabel(self.container_diario, text=f"No hay cajas cerradas para el {fecha}", font=ctk.CTkFont(size=16), text_color="#7f8c8d").grid(row=0, column=0, pady=50)
            return

        ctk.CTkLabel(self.container_diario, text="ID Caja", font=ctk.CTkFont(weight="bold"), width=80).grid(row=0, column=0, padx=5, pady=5)
        ctk.CTkLabel(self.container_diario, text="Apertura", font=ctk.CTkFont(weight="bold"), width=80).grid(row=0, column=1, padx=5, pady=5)
        ctk.CTkLabel(self.container_diario, text="Cierre", font=ctk.CTkFont(weight="bold"), width=80).grid(row=0, column=2, padx=5, pady=5)
        ctk.CTkLabel(self.container_diario, text="Cajero", font=ctk.CTkFont(weight="bold"), width=120).grid(row=0, column=3, padx=5, pady=5)
        ctk.CTkLabel(self.container_diario, text="Saldo Final Efectivo", font=ctk.CTkFont(weight="bold"), width=180).grid(row=0, column=4, padx=5, pady=5)

        for i, caja in enumerate(cajas, start=1):
            ctk.CTkLabel(self.container_diario, text=str(caja['id']), width=80).grid(row=i, column=0, padx=5, pady=5)
            ctk.CTkLabel(self.container_diario, text=caja['hora_apertura'], width=80).grid(row=i, column=1, padx=5, pady=5)
            ctk.CTkLabel(self.container_diario, text=caja['hora_cierre'] or "-", width=80).grid(row=i, column=2, padx=5, pady=5)
            ctk.CTkLabel(self.container_diario, text=caja['nombre_usuario'], width=120).grid(row=i, column=3, padx=5, pady=5)
            color = COLOR_SUCCESS if (caja['saldo_final_pesos'] or 0) >= 0 else COLOR_ERROR
            ctk.CTkLabel(self.container_diario, text=f"$ {caja['saldo_final_pesos']:,.2f}" if caja['saldo_final_pesos'] else "$ 0.00", width=180, text_color=color).grid(row=i, column=4, padx=5, pady=5)

    def _buscar_semanal(self) -> None:
        fecha = self.fecha_semanal.get()
        self._limpiar_frame(self.container_semanal)

        try:
            lunes, sabado = self.caja_controller.obtener_rango_semana(fecha)
        except ValueError:
            ctk.CTkLabel(self.container_semanal, text="Formato de fecha incorrecto. Use YYYY-MM-DD", text_color=COLOR_ERROR).grid(row=0, column=0, columnspan=3, pady=50)
            return

        ctk.CTkLabel(self.container_semanal, text=f"Resumen: {lunes} al {sabado}", font=ctk.CTkFont(size=18, weight="bold"), text_color=COLOR_SECONDARY).grid(row=0, column=0, columnspan=3, pady=(10, 30))

        resumen = self.caja_controller.obtener_resumen_por_rango(lunes, sabado)

        datos_mostrar = [
            ("🇺🇾 UYU", resumen['UYU'], COLOR_SECONDARY),
            ("💵 USD", resumen['USD'], "#27ae60"),
            ("🇧🇷 BRL", resumen['BRL'], "#8e44ad"),
        ]

        for i, (titulo, datos, color) in enumerate(datos_mostrar):
            card = ctk.CTkFrame(self.container_semanal, fg_color=COLOR_CARD, corner_radius=10)
            card.grid(row=1, column=i, padx=10, sticky="nsew")

            ctk.CTkLabel(card, text=titulo, font=ctk.CTkFont(size=16, weight="bold"), text_color=color).pack(pady=(15, 5), padx=15, anchor="w")
            ctk.CTkLabel(card, text=f"Ingresos: ${datos['ingresos']:,.2f}", font=ctk.CTkFont(size=13), text_color="#2ecc71").pack(padx=15, anchor="w")
            ctk.CTkLabel(card, text=f"Egresos: ${datos['egresos']:,.2f}", font=ctk.CTkFont(size=13), text_color="#e74c3c").pack(padx=15, anchor="w", pady=(0, 15))

        card_banco = ctk.CTkFrame(self.container_semanal, fg_color=COLOR_CARD, corner_radius=10)
        card_banco.grid(row=2, column=0, columnspan=3, padx=10, pady=(20, 10), sticky="ew")
        ctk.CTkLabel(card_banco, text=f"🏦 Total movimiento bancario (UYU): ${resumen['banco_uyu']:,.2f}", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=15, padx=15, anchor="w")

    def _buscar_mensual(self) -> None:
        self._limpiar_frame(self.container_mensual)

        try:
            fecha_str = self.fecha_mensual.get().strip()
            fecha_dt = datetime.strptime(fecha_str, "%Y-%m-%d")
            anio = fecha_dt.year
            mes = fecha_dt.month
        except ValueError:
            ctk.CTkLabel(self.container_mensual, text="Formato de fecha incorrecto. Use YYYY-MM-DD", text_color=COLOR_ERROR).pack(pady=50)
            return

        datos_semanas = self.caja_controller.obtener_datos_grafico_mensual(anio, mes)
        datos_av = self.movimiento_controller.obtener_adelantos_viaticos_mensual(anio, mes)

        if not datos_semanas and not any(d['adelantos'] > 0 or d['viaticos'] > 0 for d in datos_av.values()):
            ctk.CTkLabel(self.container_mensual, text="No hay datos para este mes.", font=ctk.CTkFont(size=16), text_color="#7f8c8d").pack(pady=50)
            return

        fig = Figure(figsize=(8, 7), dpi=100, facecolor='#2b2b2b')

        ax1 = fig.add_subplot(211)
        ax1.set_facecolor('#2b2b2b')

        if datos_semanas:
            semanas = [f"Sem {d['semana']}" for d in datos_semanas]
            ingresos = [d['ingresos_uyu'] for d in datos_semanas]
            egresos = [d['egresos_uyu'] for d in datos_semanas]

            ax1.plot(semanas, ingresos, marker='o', color='#2ecc71', linewidth=2, label='Ingresos UYU')
            ax1.plot(semanas, egresos, marker='s', color='#e74c3c', linewidth=2, label='Egresos UYU')

        ax1.set_title(f"Resumen de Caja - {mes}/{anio}", color='white', fontsize=14)
        ax1.tick_params(colors='white')
        ax1.legend(facecolor='#333333', edgecolor='white', labelcolor='white')
        ax1.grid(True, color='#444444')

        ax2 = fig.add_subplot(212)
        ax2.set_facecolor('#2b2b2b')

        semanas_keys = sorted(datos_av.keys())
        semanas_labels = [f"Sem {s}" for s in semanas_keys]
        adelantos_vals = [datos_av[s]['adelantos'] for s in semanas_keys]
        viaticos_vals = [datos_av[s]['viaticos'] for s in semanas_keys]

        x = range(len(semanas_keys))
        ancho_barra = 0.35

        ax2.bar([i - ancho_barra/2 for i in x], adelantos_vals, ancho_barra, label='Adelantos', color='#f39c12')
        ax2.bar([i + ancho_barra/2 for i in x], viaticos_vals, ancho_barra, label='Viáticos', color='#3498db')

        ax2.set_xticks(list(x))
        ax2.set_xticklabels(semanas_labels)
        ax2.set_title(f"Adelantos y Viáticos - {mes}/{anio}", color='white', fontsize=14)
        ax2.tick_params(colors='white')
        ax2.legend(facecolor='#333333', edgecolor='white', labelcolor='white')
        ax2.grid(True, color='#444444', axis='y')

        fig.tight_layout(pad=3.0)

        canvas = FigureCanvasTkAgg(fig, master=self.container_mensual)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    # --- LÓGICA ADELANTOS Y VIÁTICOS ---

    def _buscar_adelantos_viaticos(self) -> None:
        desde = self.fecha_desde_av.get().strip()
        hasta = self.fecha_hasta_av.get().strip()

        try:
            datetime.strptime(desde, "%Y-%m-%d")
            datetime.strptime(hasta, "%Y-%m-%d")
        except ValueError:
            ctk.CTkLabel(
                self.container_av, text="❌ Formato de fecha inválido. Use YYYY-MM-DD",
                font=ctk.CTkFont(size=14), text_color=COLOR_ERROR
            ).grid(row=0, column=0, columnspan=4, pady=20, padx=10)
            return

        self._limpiar_frame(self.container_av)

        adelantos = self.movimiento_controller.obtener_adelantos_por_rango(desde, hasta)
        viaticos = self.movimiento_controller.obtener_viaticos_por_rango(desde, hasta)

        self._ultimos_adelantos = adelantos
        if adelantos:
            self.btn_excel_adelantos.configure(state="normal")
        else:
            self.btn_excel_adelantos.configure(state="disabled")

        if not adelantos and not viaticos:
            ctk.CTkLabel(
                self.container_av, text="No se encontraron adelantos ni viáticos en ese rango.",
                font=ctk.CTkFont(size=14), text_color="#7f8c8d"
            ).grid(row=0, column=0, columnspan=4, pady=20, padx=10)
            return

        row_actual = 0

        ctk.CTkLabel(
            self.container_av, text="💰 Adelantos de Sueldo",
            font=ctk.CTkFont(size=16, weight="bold"), text_color="#f39c12"
        ).grid(row=row_actual, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5))
        row_actual += 1

        if not adelantos:
            ctk.CTkLabel(
                self.container_av, text="Sin adelantos en este periodo.",
                font=ctk.CTkFont(size=12), text_color="#7f8c8d"
            ).grid(row=row_actual, column=0, columnspan=4, sticky="w", padx=10, pady=3)
            row_actual += 1
            total_general_adelantos = 0.0
        else:
            agrupados = defaultdict(list)
            for adv in adelantos:
                nombre = adv.get('empleado') or 'SIN ASIGNAR'
                agrupados[nombre].append(adv)

            total_general_adelantos = 0.0

            for nombre_emp, lista_adelantos in agrupados.items():
                ctk.CTkLabel(
                    self.container_av, text=f"👤 {nombre_emp}",
                    font=ctk.CTkFont(size=13, weight="bold"), text_color=COLOR_TEXT
                ).grid(row=row_actual, column=0, columnspan=4, sticky="w", padx=(20, 0), pady=(10, 3))
                row_actual += 1

                sub_headers = ["Fecha", "Motivo", "Monto", "Moneda"]
                sub_widths = [150, 250, 120, 80]
                for col, (h, w) in enumerate(zip(sub_headers, sub_widths)):
                    ctk.CTkLabel(
                        self.container_av, text=h, font=ctk.CTkFont(size=11, weight="bold"),
                        text_color="#95a5a6", width=w, anchor="w"
                    ).grid(row=row_actual, column=col, padx=(30, 3), pady=2)
                row_actual += 1

                subtotal_emp = 0.0
                for adv in lista_adelantos:
                    fecha = adv['created_at'].split(' ')[0] if ' ' in adv['created_at'] else adv['created_at']
                    vals = [fecha, adv['motivo'], f"${adv['monto']:,.2f}", adv['moneda']]
                    subtotal_emp += adv['monto'] if adv['moneda'] == 'UYU' else 0

                    for col, (v, w) in enumerate(zip(vals, sub_widths)):
                        ctk.CTkLabel(
                            self.container_av, text=v, font=ctk.CTkFont(size=12),
                            width=w, anchor="w"
                        ).grid(row=row_actual, column=col, padx=(30, 3), pady=1)
                    row_actual += 1

                ctk.CTkLabel(
                    self.container_av, text=f"Subtotal {nombre_emp}: ${subtotal_emp:,.2f}",
                    font=ctk.CTkFont(size=12, weight="bold"), text_color="#f39c12"
                ).grid(row=row_actual, column=0, columnspan=4, sticky="w", padx=(30, 0), pady=(3, 5))
                row_actual += 1

                total_general_adelantos += subtotal_emp

            ctk.CTkLabel(
                self.container_av, text=f"TOTAL ADELANTOS (UYU): ${total_general_adelantos:,.2f}",
                font=ctk.CTkFont(size=14, weight="bold"), text_color="#f39c12"
            ).grid(row=row_actual, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 20))
            row_actual += 1

        ctk.CTkLabel(
            self.container_av, text="🚗 Viáticos",
            font=ctk.CTkFont(size=16, weight="bold"), text_color="#3498db"
        ).grid(row=row_actual, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5))
        row_actual += 1

        headers_viaticos = ["Fecha", "Motivo", "Monto", "Moneda"]
        widths_viaticos = [150, 250, 120, 80]
        for col, (h, w) in enumerate(zip(headers_viaticos, widths_viaticos)):
            ctk.CTkLabel(
                self.container_av, text=h, font=ctk.CTkFont(weight="bold", size=12),
                text_color=COLOR_SECONDARY, width=w, anchor="w"
            ).grid(row=row_actual, column=col, padx=3, pady=3)
        row_actual += 1

        total_viaticos = 0.0
        if not viaticos:
            ctk.CTkLabel(
                self.container_av, text="Sin viáticos en este periodo.",
                font=ctk.CTkFont(size=12), text_color="#7f8c8d"
            ).grid(row=row_actual, column=0, columnspan=4, sticky="w", padx=10, pady=3)
            row_actual += 1
        else:
            for viat in viaticos:
                fecha = viat['created_at'].split(' ')[0] if ' ' in viat['created_at'] else viat['created_at']
                vals = [fecha, viat['motivo'], f"${viat['monto']:,.2f}", viat['moneda']]
                total_viaticos += viat['monto'] if viat['moneda'] == 'UYU' else 0
                for col, (v, w) in enumerate(zip(vals, widths_viaticos)):
                    ctk.CTkLabel(self.container_av, text=v, font=ctk.CTkFont(size=12), width=w, anchor="w").grid(
                        row=row_actual, column=col, padx=3, pady=2
                    )
                row_actual += 1

        ctk.CTkLabel(
            self.container_av, text=f"TOTAL VIÁTICOS (UYU): ${total_viaticos:,.2f}",
            font=ctk.CTkFont(size=13, weight="bold"), text_color="#3498db"
        ).grid(row=row_actual, column=0, columnspan=4, sticky="w", padx=10, pady=(5, 10))
        row_actual += 1

        ctk.CTkLabel(
            self.container_av, text=f"📊 TOTAL GENERAL (UYU): ${total_general_adelantos + total_viaticos:,.2f}",
            font=ctk.CTkFont(size=15, weight="bold"), text_color=COLOR_TEXT
        ).grid(row=row_actual, column=0, columnspan=4, sticky="w", padx=10, pady=(15, 10))

    # --- EXPORTAR EXCEL ---

    def _exportar_adelantos_excel(self) -> None:
        if not self._ultimos_adelantos:
            return

        desde = self.fecha_desde_av.get().strip()
        hasta = self.fecha_hasta_av.get().strip()

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            initialfile=f"Adelantos_{desde}_a_{hasta}.xlsx",
            title="Guardar Reporte de Adelantos"
        )

        if not filepath:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Adelantos"

            font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_emp = Font(name="Calibri", size=11, bold=True, color="1F4E79")
            font_normal = Font(name="Calibri", size=11)
            font_subtotal = Font(name="Calibri", size=11, bold=True, color="E67E22")
            font_total = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

            fill_title = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
            fill_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            fill_subtotal = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
            fill_total = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")

            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            row = 1
            ws.merge_cells('A1:D1')
            ws['A1'] = f"Reporte de Adelantos: {desde} a {hasta}"
            ws['A1'].font = font_title
            ws['A1'].fill = fill_title
            ws['A1'].alignment = Alignment(horizontal="center")
            row += 2

            agrupados = defaultdict(list)
            for adv in self._ultimos_adelantos:
                nombre = adv.get('empleado') or 'SIN ASIGNAR'
                agrupados[nombre].append(adv)

            total_general = 0.0

            for nombre_emp, lista in agrupados.items():
                ws[f'A{row}'] = f"Empleado: {nombre_emp}"
                ws[f'A{row}'].font = font_emp
                row += 1

                headers = ["Fecha", "Motivo", "Monto", "Moneda"]
                for col, h in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col, value=h)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                row += 1

                subtotal = 0.0
                for adv in lista:
                    fecha = adv['created_at'].split(' ')[0] if ' ' in adv['created_at'] else adv['created_at']
                    monto = adv['monto']
                    vals = [fecha, adv['motivo'], monto, adv['moneda']]
                    if adv['moneda'] == 'UYU':
                        subtotal += monto

                    for col, v in enumerate(vals, start=1):
                        cell = ws.cell(row=row, column=col, value=v)
                        cell.font = font_normal
                        cell.border = thin_border
                        if col == 3:
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal="right")
                    row += 1

                ws.merge_cells(f'A{row}:C{row}')
                ws[f'A{row}'] = f"Subtotal {nombre_emp}"
                ws[f'A{row}'].font = font_subtotal
                ws[f'A{row}'].fill = fill_subtotal
                ws[f'A{row}'].border = thin_border
                ws[f'D{row}'] = subtotal
                ws[f'D{row}'].font = font_subtotal
                ws[f'D{row}'].fill = fill_subtotal
                ws[f'D{row}'].border = thin_border
                ws[f'D{row}'].number_format = '#,##0.00'
                ws[f'D{row}'].alignment = Alignment(horizontal="right")
                row += 1
                total_general += subtotal

            row += 1

            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = "TOTAL GENERAL ADELANTOS (UYU)"
            ws[f'A{row}'].font = font_total
            ws[f'A{row}'].fill = fill_total
            ws[f'A{row}'].border = thin_border
            ws[f'D{row}'] = total_general
            ws[f'D{row}'].font = font_total
            ws[f'D{row}'].fill = fill_total
            ws[f'D{row}'].border = thin_border
            ws[f'D{row}'].number_format = '#,##0.00'
            ws[f'D{row}'].alignment = Alignment(horizontal="right")

            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12

            wb.save(filepath)

            ctk.CTkLabel(
                self.container_av,
                text=f"✅ Excel guardado: {filepath}",
                font=ctk.CTkFont(size=12), text_color=COLOR_SUCCESS
            ).grid(row=100, column=0, columnspan=4, pady=5, sticky="w", padx=10)

        except Exception as e:
            ctk.CTkLabel(
                self.container_av,
                text=f"❌ Error al generar Excel: {str(e)}",
                font=ctk.CTkFont(size=12), text_color=COLOR_ERROR
            ).grid(row=100, column=0, columnspan=4, pady=5, sticky="w", padx=10)
