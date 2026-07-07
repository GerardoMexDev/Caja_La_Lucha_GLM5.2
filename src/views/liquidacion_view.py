"""
Vista de Liquidación de Sueldos Semanales.
"""

import customtkinter as ctk
from datetime import datetime, timedelta
from typing import Dict, Any, List
from tkinter import Toplevel, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from tkcalendar import Calendar

from controllers.empleado_controller import EmpleadoController
from utils.constants import COLOR_CARD, COLOR_TEXT, COLOR_SECONDARY, COLOR_SUCCESS, COLOR_ERROR, COLOR_ACCENT


class LiquidacionFrame(ctk.CTkFrame):
    def __init__(self, parent, empleado_controller: EmpleadoController, on_back) -> None:
        super().__init__(parent, fg_color="transparent")
        self.empleado_controller = empleado_controller
        self.on_back = on_back
        self._ultima_liquidacion: List[Dict[str, Any]] = []
        self._create_widgets()

    def _create_widgets(self) -> None:
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", pady=(0, 20))
        
        ctk.CTkButton(
            header, text="← Volver", width=100, height=35,
            fg_color="transparent", text_color=COLOR_TEXT,
            border_width=1, border_color="#555",
            command=self.on_back
        ).pack(side="left")

        ctk.CTkLabel(
            header, text="💵 Liquidación de Sueldos", font=ctk.CTkFont(size=22, weight="bold"),
            text_color=COLOR_TEXT
        ).pack(side="left", padx=20)

        # Filtros
        filtros_card = ctk.CTkFrame(self, fg_color=COLOR_CARD, corner_radius=15)
        filtros_card.pack(fill="x", pady=(0, 20))

        filtros_grid = ctk.CTkFrame(filtros_card, fg_color="transparent")
        filtros_grid.pack(fill="x", padx=20, pady=15)

        ctk.CTkLabel(filtros_grid, text="Desde:", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=(0, 5))
        self.fecha_desde = ctk.CTkEntry(filtros_grid, placeholder_text="YYYY-MM-DD", width=150)
        self.fecha_desde.grid(row=0, column=1, padx=(0, 15))
        self.fecha_desde.insert(0, (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d"))
        self.fecha_desde.configure(cursor="hand2")
        self.fecha_desde.bind("<Button-1>", lambda e: self._abrir_calendario(self.fecha_desde))

        ctk.CTkLabel(filtros_grid, text="Hasta:", font=ctk.CTkFont(size=14)).grid(row=0, column=2, padx=(0, 5))
        self.fecha_hasta = ctk.CTkEntry(filtros_grid, placeholder_text="YYYY-MM-DD", width=150)
        self.fecha_hasta.grid(row=0, column=3, padx=(0, 15))
        self.fecha_hasta.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.fecha_hasta.configure(cursor="hand2")
        self.fecha_hasta.bind("<Button-1>", lambda e: self._abrir_calendario(self.fecha_hasta))

        ctk.CTkButton(filtros_grid, text="🔍 Calcular Liquidación", width=200, height=35, fg_color=COLOR_SECONDARY, hover_color=COLOR_ACCENT, command=self._calcular).grid(row=0, column=4, padx=(10, 0))

        # Botón Excel
        self.btn_excel = ctk.CTkButton(
            filtros_grid, text="📥 Exportar a Excel", width=180, height=35,
            fg_color="#1e8449", hover_color="#196f3d", command=self._exportar_excel, state="disabled"
        )
        self.btn_excel.grid(row=1, column=0, columnspan=5, sticky="e", pady=(15, 0))

        # Tabla
        table_card = ctk.CTkFrame(self, fg_color=COLOR_CARD, corner_radius=15)
        table_card.pack(fill="both", expand=True)

        self.table_container = ctk.CTkScrollableFrame(table_card, fg_color="transparent")
        self.table_container.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(self.table_container, text="Seleccione un rango de fechas y presione 'Calcular Liquidación'.", text_color="#7f8c8d", font=ctk.CTkFont(size=14)).pack(pady=40)

    def _abrir_calendario(self, entry_widget) -> None:
        top = Toplevel(self.winfo_toplevel())
        top.title("Seleccionar Fecha")
        top.configure(bg="#2b2b2b")
        top.resizable(False, False)
        top.grab_set()
        top.geometry("+%d+%d" % (self.winfo_rootx() + 50, self.winfo_rooty() + 50))

        cal = Calendar(top, selectmode="day", date_pattern="yyyy-mm-dd", background="#2b2b2b", foreground="white", selectbackground="#3498db", selectforeground="white", borderwidth=0, headersbackground="#34495e")

        def seleccionar_fecha(event):
            fecha_str = cal.get_date()
            entry_widget.delete(0, "end")
            entry_widget.insert(0, fecha_str)
            top.destroy()

        cal.bind("<<CalendarSelected>>", seleccionar_fecha)
        cal.pack(padx=10, pady=10)
        ctk.CTkButton(top, text="Cancelar", width=100, fg_color="#7f8c8d", command=top.destroy).pack(pady=(0, 10))

    def _calcular(self) -> None:
        desde = self.fecha_desde.get().strip()
        hasta = self.fecha_hasta.get().strip()

        try:
            datetime.strptime(desde, "%Y-%m-%d")
            datetime.strptime(hasta, "%Y-%m-%d")
        except ValueError:
            self._limpiar_tabla()
            ctk.CTkLabel(self.table_container, text="❌ Formato de fecha inválido.", text_color=COLOR_ERROR, font=ctk.CTkFont(size=14)).pack(pady=20)
            return

        self._limpiar_tabla()
        datos = self.empleado_controller.calcular_liquidacion_por_rango(desde, hasta)
        self._ultima_liquidacion = datos

        if datos:
            self.btn_excel.configure(state="normal")
        else:
            self.btn_excel.configure(state="disabled")

        if not datos:
            ctk.CTkLabel(self.table_container, text="No hay empleados activos para liquidar.", text_color="#7f8c8d", font=ctk.CTkFont(size=14)).pack(pady=20)
            return

        ctk.CTkLabel(self.table_container, text=f"Período: {desde} al {hasta}", font=ctk.CTkFont(size=14, weight="bold"), text_color=COLOR_ACCENT).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 15))

        headers = [("Empleado", 250), ("Sueldo Base", 150), ("Adelantos (UYU)", 150), ("A Cobrar (UYU)", 150)]
        for col, (h, w) in enumerate(headers):
            ctk.CTkLabel(self.table_container, text=h, font=ctk.CTkFont(weight="bold", size=13), text_color=COLOR_SECONDARY, width=w, anchor="w").grid(row=1, column=col, padx=5, pady=5)

        total_base = 0.0
        total_adelantos = 0.0
        total_a_cobrar = 0.0

        for i, row in enumerate(datos, start=2):
            color_pago = COLOR_SUCCESS if row['pago_final'] >= 0 else COLOR_ERROR
            signo = "" if row['pago_final'] >= 0 else "- "

            ctk.CTkLabel(self.table_container, text=row['nombre'], width=250, anchor="w").grid(row=i, column=0, padx=5, pady=5)
            ctk.CTkLabel(self.table_container, text=f"$ {row['sueldo_base']:,.2f}", width=150, anchor="w").grid(row=i, column=1, padx=5, pady=5)
            ctk.CTkLabel(self.table_container, text=f"- $ {row['total_adelantos']:,.2f}", width=150, anchor="w", text_color="#e74c3c").grid(row=i, column=2, padx=5, pady=5)
            ctk.CTkLabel(self.table_container, text=f"{signo}$ {abs(row['pago_final']):,.2f}", width=150, anchor="w", font=ctk.CTkFont(weight="bold"), text_color=color_pago).grid(row=i, column=3, padx=5, pady=5)

            total_base += row['sueldo_base']
            total_adelantos += row['total_adelantos']
            total_a_cobrar += row['pago_final']

        fila_total = len(datos) + 2
        ctk.CTkLabel(self.table_container, text="TOTALES", font=ctk.CTkFont(weight="bold", size=13), width=250, anchor="w").grid(row=fila_total, column=0, padx=5, pady=(15, 5))
        ctk.CTkLabel(self.table_container, text=f"$ {total_base:,.2f}", font=ctk.CTkFont(weight="bold"), width=150, anchor="w").grid(row=fila_total, column=1, padx=5, pady=(15, 5))
        ctk.CTkLabel(self.table_container, text=f"- $ {total_adelantos:,.2f}", font=ctk.CTkFont(weight="bold"), width=150, anchor="w", text_color="#e74c3c").grid(row=fila_total, column=2, padx=5, pady=(15, 5))
        ctk.CTkLabel(self.table_container, text=f"$ {total_a_cobrar:,.2f}", font=ctk.CTkFont(weight="bold", size=14), width=150, anchor="w", text_color=COLOR_ACCENT).grid(row=fila_total, column=3, padx=5, pady=(15, 5))

    def _exportar_excel(self) -> None:
        if not self._ultima_liquidacion:
            return

        desde = self.fecha_desde.get().strip()
        hasta = self.fecha_hasta.get().strip()
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")], initialfile=f"Liquidacion_{desde}_a_{hasta}.xlsx")

        if not filepath:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Liquidación"

            font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_normal = Font(name="Calibri", size=11)
            font_total = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

            fill_title = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
            fill_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            fill_total = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            ws.merge_cells('A1:D1')
            ws['A1'] = f"Liquidación de Sueldos: {desde} a {hasta}"
            ws['A1'].font = font_title
            ws['A1'].fill = fill_title
            ws['A1'].alignment = Alignment(horizontal="center")

            headers = ["Empleado", "Sueldo Base", "Adelantos", "A Cobrar"]
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.border = thin_border

            total_base = total_adelantos = total_a_cobrar = 0.0
            for i, row in enumerate(self._ultima_liquidacion, start=4):
                ws.cell(row=i, column=1, value=row['nombre']).border = thin_border
                ws.cell(row=i, column=2, value=row['sueldo_base']).border = thin_border
                ws.cell(row=i, column=2).number_format = '#,##0.00'
                ws.cell(row=i, column=3, value=row['total_adelantos']).border = thin_border
                ws.cell(row=i, column=3).number_format = '#,##0.00'
                ws.cell(row=i, column=4, value=row['pago_final']).border = thin_border
                ws.cell(row=i, column=4).number_format = '#,##0.00'
                if row['pago_final'] < 0:
                    ws.cell(row=i, column=4).font = Font(color="FF0000", bold=True)

                total_base += row['sueldo_base']
                total_adelantos += row['total_adelantos']
                total_a_cobrar += row['pago_final']

            fila_total = 4 + len(self._ultima_liquidacion)
            for col, val in enumerate([None, total_base, total_adelantos, total_a_cobrar], start=1):
                cell = ws.cell(row=fila_total, column=col, value=val if val is not None else "TOTALES")
                cell.font = font_total
                cell.fill = fill_total
                cell.border = thin_border
                if col > 1:
                    cell.number_format = '#,##0.00'

            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18

            wb.save(filepath)
            ctk.CTkLabel(self.table_container, text=f"✅ Excel guardado correctamente.", font=ctk.CTkFont(size=12), text_color=COLOR_SUCCESS).grid(row=100, column=0, columnspan=4, pady=10)
        except Exception as e:
            ctk.CTkLabel(self.table_container, text=f"❌ Error al guardar: {str(e)}", font=ctk.CTkFont(size=12), text_color=COLOR_ERROR).grid(row=100, column=0, columnspan=4, pady=10)

    def _limpiar_tabla(self) -> None:
        for widget in self.table_container.winfo_children():
            widget.destroy()