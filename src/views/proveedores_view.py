"""
Vista de Gestión de Proveedores y Cuentas por Pagar.
"""

import customtkinter as ctk
from datetime import datetime
from typing import Dict, Any, Callable, List, Optional

from controllers.proveedor_controller import ProveedorController
from utils.constants import (
    COLOR_PRIMARY, COLOR_SECONDARY, COLOR_ACCENT, COLOR_BG,
    COLOR_CARD, COLOR_TEXT, COLOR_ERROR
)


class ProveedoresFrame(ctk.CTkFrame):
    """Frame con dos pestañas: Proveedores y Facturas."""

    def __init__(
        self,
        parent,
        proveedor_controller: ProveedorController,
        on_back: Callable[[], None],
    ) -> None:
        super().__init__(parent, fg_color=COLOR_BG)
        self.proveedor_controller = proveedor_controller
        self.on_back = on_back
        self._create_widgets()
        self._cargar_datos()

    # ── WIDGETS ──────────────────────────────────────────────────────────────

    def _create_widgets(self) -> None:
        title_frame = ctk.CTkFrame(self, fg_color="transparent")
        title_frame.grid(row=0, column=0, sticky="ew", padx=20, pady=(15, 10))
        title_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            title_frame, text="📦 Proveedores y Cuentas por Pagar",
            font=ctk.CTkFont(size=22, weight="bold"), text_color=COLOR_TEXT
        ).grid(row=0, column=0, sticky="w")

        ctk.CTkButton(
            title_frame, text="← Volver", width=100, height=35,
            fg_color=COLOR_SECONDARY, hover_color=COLOR_PRIMARY,
            command=self.on_back
        ).grid(row=0, column=1, sticky="e")

        self.tabview = ctk.CTkTabview(self, fg_color=COLOR_CARD)
        self.tabview.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 20))
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._create_tab_proveedores()
        self._create_tab_facturas()

    # ── TAB PROVEEDORES ──────────────────────────────────────────────────────

    def _create_tab_proveedores(self) -> None:
        tab = self.tabview.add("📋 Proveedores")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(1, weight=1)

        form = ctk.CTkFrame(tab, fg_color=COLOR_BG, corner_radius=10)
        form.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        for i in range(4):
            form.grid_columnconfigure(i, weight=1)

        labels = ["Nombre *", "Teléfono", "Email", "Dirección"]
        placeholders = ["Nombre del proveedor", "099123456", "correo@ejemplo.com", "Dirección"]
        self._prov_entries: List[ctk.CTkEntry] = []

        for col, (lbl, ph) in enumerate(zip(labels, placeholders)):
            ctk.CTkLabel(form, text=lbl, text_color=COLOR_TEXT, font=ctk.CTkFont(size=12)).grid(
                row=0, column=col, sticky="w", padx=10, pady=(10, 0)
            )
            entry = ctk.CTkEntry(form, placeholder_text=ph, height=38)
            entry.grid(row=1, column=col, sticky="ew", padx=10, pady=(0, 10))
            self._prov_entries.append(entry)

        ctk.CTkButton(
            form, text="✅ Guardar", height=38, width=120,
            fg_color=COLOR_ACCENT, hover_color=COLOR_PRIMARY,
            command=self._guardar_proveedor
        ).grid(row=1, column=4, padx=10, pady=(0, 10))

        self.prov_table = ctk.CTkScrollableFrame(tab, fg_color=COLOR_CARD, corner_radius=10)
        self.prov_table.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        for i in range(5):
            self.prov_table.grid_columnconfigure(i, weight=1)

        headers = ["ID", "Nombre", "Teléfono", "Email", "Estado"]
        for col, h in enumerate(headers):
            ctk.CTkLabel(
                self.prov_table, text=h,
                font=ctk.CTkFont(size=12, weight="bold"), text_color=COLOR_ACCENT
            ).grid(row=0, column=col, padx=10, pady=8, sticky="w")

    # ── TAB FACTURAS ─────────────────────────────────────────────────────────

    def _create_tab_facturas(self) -> None:
        tab = self.tabview.add("📄 Facturas")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(2, weight=1)

        # Tarjetas de resumen
        resumen = ctk.CTkFrame(tab, fg_color="transparent")
        resumen.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        resumen.grid_columnconfigure((0, 1, 2, 3), weight=1)

        c1 = ctk.CTkFrame(resumen, fg_color=COLOR_CARD, corner_radius=10, height=70)
        c1.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        c1.grid_propagate(False)
        ctk.CTkLabel(c1, text="Facturas Pendientes", font=ctk.CTkFont(size=11), text_color="#bdc3c7").pack(pady=(12, 0))
        self.lbl_cant_pendientes = ctk.CTkLabel(c1, text="0", font=ctk.CTkFont(size=22, weight="bold"), text_color=COLOR_ACCENT)
        self.lbl_cant_pendientes.pack()

        c2 = ctk.CTkFrame(resumen, fg_color=COLOR_CARD, corner_radius=10, height=70)
        c2.grid(row=0, column=1, sticky="ew", padx=5)
        c2.grid_propagate(False)
        ctk.CTkLabel(c2, text="Total UYU", font=ctk.CTkFont(size=11), text_color="#bdc3c7").pack(pady=(12, 0))
        self.lbl_total_uyu = ctk.CTkLabel(c2, text="$ 0.00", font=ctk.CTkFont(size=20, weight="bold"), text_color=COLOR_ERROR)
        self.lbl_total_uyu.pack()

        c3 = ctk.CTkFrame(resumen, fg_color=COLOR_CARD, corner_radius=10, height=70)
        c3.grid(row=0, column=2, sticky="ew", padx=5)
        c3.grid_propagate(False)
        ctk.CTkLabel(c3, text="Total USD", font=ctk.CTkFont(size=11), text_color="#bdc3c7").pack(pady=(12, 0))
        self.lbl_total_usd = ctk.CTkLabel(c3, text="U$S 0.00", font=ctk.CTkFont(size=20, weight="bold"), text_color="#27ae60")
        self.lbl_total_usd.pack()

        c4 = ctk.CTkFrame(resumen, fg_color=COLOR_CARD, corner_radius=10, height=70)
        c4.grid(row=0, column=3, sticky="ew", padx=(5, 0))
        c4.grid_propagate(False)
        ctk.CTkButton(
            c4, text="📊 Exportar Excel", height=38,
            fg_color="#27ae60", hover_color="#2ecc71",
            command=self._exportar_facturas_excel
        ).pack(expand=True)

        # Formulario de factura
        form = ctk.CTkFrame(tab, fg_color=COLOR_BG, corner_radius=10)
        form.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))
        for i in range(5):
            form.grid_columnconfigure(i, weight=1)

        # Fila 1
        ctk.CTkLabel(form, text="Proveedor *", text_color=COLOR_TEXT, font=ctk.CTkFont(size=12)).grid(row=0, column=0, sticky="w", padx=10, pady=(10, 0))
        self.combo_proveedor = ctk.CTkComboBox(form, height=38)
        self.combo_proveedor.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 5))

        ctk.CTkLabel(form, text="Descripción *", text_color=COLOR_TEXT, font=ctk.CTkFont(size=12)).grid(row=0, column=1, sticky="w", padx=10, pady=(10, 0))
        self.entry_fact_desc = ctk.CTkEntry(form, placeholder_text="Concepto de la factura", height=38)
        self.entry_fact_desc.grid(row=1, column=1, sticky="ew", padx=10, pady=(0, 5))

        ctk.CTkLabel(form, text="Fecha Compra *", text_color=COLOR_TEXT, font=ctk.CTkFont(size=12)).grid(row=0, column=2, sticky="w", padx=10, pady=(10, 0))
        self.entry_fact_fecha = ctk.CTkEntry(form, placeholder_text="YYYY-MM-DD", height=38)
        self.entry_fact_fecha.grid(row=1, column=2, sticky="ew", padx=10, pady=(0, 5))
        self.entry_fact_fecha.insert(0, datetime.now().strftime("%Y-%m-%d"))

        ctk.CTkLabel(form, text="Monto Base *", text_color=COLOR_TEXT, font=ctk.CTkFont(size=12)).grid(row=0, column=3, sticky="w", padx=10, pady=(10, 0))
        self.entry_fact_monto = ctk.CTkEntry(form, placeholder_text="0.00", height=38)
        self.entry_fact_monto.grid(row=1, column=3, sticky="ew", padx=10, pady=(0, 5))
        self.entry_fact_monto.bind("<KeyRelease>", self._recalcular)

        ctk.CTkLabel(form, text="Moneda", text_color=COLOR_TEXT, font=ctk.CTkFont(size=12)).grid(row=0, column=4, sticky="w", padx=10, pady=(10, 0))
        self.combo_moneda = ctk.CTkComboBox(form, values=["UYU", "USD"], height=38, width=90)
        self.combo_moneda.grid(row=1, column=4, sticky="ew", padx=10, pady=(0, 5))
        self.combo_moneda.set("UYU")

        # Fila 2
        ctk.CTkLabel(form, text="IVA %", text_color=COLOR_TEXT, font=ctk.CTkFont(size=12)).grid(row=2, column=0, sticky="w", padx=10, pady=(5, 0))
        self.entry_fact_iva_pct = ctk.CTkEntry(form, placeholder_text="22", height=38)
        self.entry_fact_iva_pct.grid(row=3, column=0, sticky="ew", padx=10, pady=(0, 5))
        self.entry_fact_iva_pct.insert(0, "22")
        self.entry_fact_iva_pct.bind("<KeyRelease>", self._recalcular)

        ctk.CTkLabel(form, text="IVA Monto", text_color="#bdc3c7", font=ctk.CTkFont(size=12)).grid(row=2, column=1, sticky="w", padx=10, pady=(5, 0))
        self.lbl_iva_monto = ctk.CTkLabel(form, text="$ 0.00", font=ctk.CTkFont(size=15, weight="bold"), text_color=COLOR_TEXT)
        self.lbl_iva_monto.grid(row=3, column=1, sticky="w", padx=10, pady=(0, 5))

        ctk.CTkLabel(form, text="TOTAL", text_color=COLOR_ACCENT, font=ctk.CTkFont(size=13, weight="bold")).grid(row=2, column=2, sticky="w", padx=10, pady=(5, 0))
        self.lbl_fact_total = ctk.CTkLabel(form, text="$ 0.00", font=ctk.CTkFont(size=18, weight="bold"), text_color=COLOR_ACCENT)
        self.lbl_fact_total.grid(row=3, column=2, sticky="w", padx=10, pady=(0, 5))

        ctk.CTkLabel(form, text="Fecha Pago Est.", text_color=COLOR_TEXT, font=ctk.CTkFont(size=12)).grid(row=2, column=3, sticky="w", padx=10, pady=(5, 0))
        self.entry_fact_fecha_pago = ctk.CTkEntry(form, placeholder_text="YYYY-MM-DD (opcional)", height=38)
        self.entry_fact_fecha_pago.grid(row=3, column=3, sticky="ew", padx=10, pady=(0, 5))

        ctk.CTkLabel(form, text="Observaciones", text_color=COLOR_TEXT, font=ctk.CTkFont(size=12)).grid(row=2, column=4, sticky="w", padx=10, pady=(5, 0))
        self.entry_fact_obs = ctk.CTkEntry(form, placeholder_text="Opcional", height=38)
        self.entry_fact_obs.grid(row=3, column=4, sticky="ew", padx=10, pady=(0, 5))

        # Fila 3: Botón registrar
        ctk.CTkButton(
            form, text="📄 Registrar Factura", height=38, width=180,
            fg_color=COLOR_ACCENT, hover_color=COLOR_PRIMARY,
            command=self._guardar_factura
        ).grid(row=4, column=0, columnspan=5, padx=10, pady=(5, 10), sticky="e")

        # Filtro + Tabla
        filtro_frame = ctk.CTkFrame(tab, fg_color="transparent")
        filtro_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 10))
        filtro_frame.grid_columnconfigure(0, weight=1)
        filtro_frame.grid_rowconfigure(1, weight=1)

        barra = ctk.CTkFrame(filtro_frame, fg_color="transparent")
        barra.grid(row=0, column=0, sticky="ew", pady=(0, 5))

        ctk.CTkLabel(barra, text="Filtrar:", text_color=COLOR_TEXT, font=ctk.CTkFont(size=12)).pack(side="left", padx=(0, 8))
        self.combo_filtro = ctk.CTkComboBox(
            barra, values=["Todas", "Pendiente", "Pagada", "Cancelada"],
            width=140, height=30, command=self._filtrar_facturas
        )
        self.combo_filtro.set("Todas")
        self.combo_filtro.pack(side="left")

        self.fact_table = ctk.CTkScrollableFrame(filtro_frame, fg_color=COLOR_CARD, corner_radius=10)
        self.fact_table.grid(row=1, column=0, sticky="nsew")
        self.fact_table.grid_columnconfigure((1, 2), weight=2)
        self.fact_table.grid_columnconfigure((0, 3, 4, 5, 6, 7, 8), weight=1)

        headers = ["ID", "Proveedor", "Descripción", "Moneda", "Fecha", "Base", "IVA", "Total", "Estado", "Acciones"]
        for col, h in enumerate(headers):
            ctk.CTkLabel(
                self.fact_table, text=h,
                font=ctk.CTkFont(size=11, weight="bold"), text_color=COLOR_ACCENT
            ).grid(row=0, column=col, padx=6, pady=8, sticky="w")

    # ── CARGA DE DATOS ───────────────────────────────────────────────────────

    def _cargar_datos(self) -> None:
        self._render_proveedores()
        self._actualizar_combo_proveedores()
        self._actualizar_resumen()
        self._render_facturas()

    def _render_proveedores(self) -> None:
        for widget in self.prov_table.winfo_children():
            info = widget.grid_info()
            if int(info.get("row", 0)) >= 1:
                widget.destroy()

        proveedores = self.proveedor_controller.obtener_proveedores(activos_solo=False)
        for i, p in enumerate(proveedores, start=1):
            estado_text = "✅ Activo" if p["activo"] == 1 else "❌ Inactivo"
            estado_color = "#2ecc71" if p["activo"] == 1 else "#e74c3c"

            ctk.CTkLabel(self.prov_table, text=str(p["id"]), text_color=COLOR_TEXT, font=ctk.CTkFont(size=11)).grid(row=i, column=0, padx=10, pady=4, sticky="w")
            ctk.CTkLabel(self.prov_table, text=p["nombre"], text_color=COLOR_TEXT, font=ctk.CTkFont(size=11)).grid(row=i, column=1, padx=10, pady=4, sticky="w")
            ctk.CTkLabel(self.prov_table, text=p["telefono"] or "-", text_color="#bdc3c7", font=ctk.CTkFont(size=11)).grid(row=i, column=2, padx=10, pady=4, sticky="w")
            ctk.CTkLabel(self.prov_table, text=p["email"] or "-", text_color="#bdc3c7", font=ctk.CTkFont(size=11)).grid(row=i, column=3, padx=10, pady=4, sticky="w")

            btn_estado = ctk.CTkButton(
                self.prov_table, text=estado_text, width=100, height=28,
                font=ctk.CTkFont(size=10),
                fg_color="transparent", text_color=estado_color,
                border_width=1, border_color=estado_color,
                hover_color="#1c2833",
                command=lambda pid=p["id"]: self._toggle_proveedor(pid)
            )
            btn_estado.grid(row=i, column=4, padx=10, pady=4, sticky="w")

    def _render_facturas(self, estado_filtro: Optional[str] = None) -> None:
        for widget in self.fact_table.winfo_children():
            info = widget.grid_info()
            if int(info.get("row", 0)) >= 1:
                widget.destroy()

        facturas = self.proveedor_controller.obtener_facturas(estado=estado_filtro)
        colores_estado = {
            "pendiente": ("#f39c12", "#1c2833"),
            "pagada": ("#2ecc71", "#1c2833"),
            "cancelada": ("#e74c3c", "#1c2833"),
        }

        for i, f in enumerate(facturas, start=1):
            fg, bg = colores_estado.get(f["estado"], (COLOR_TEXT, COLOR_CARD))
            moneda = f.get("moneda", "UYU") or "UYU"
            simbolo = "U$S" if moneda == "USD" else "$"

            ctk.CTkLabel(self.fact_table, text=str(f["id"]), text_color="#bdc3c7", font=ctk.CTkFont(size=11)).grid(row=i, column=0, padx=6, pady=4, sticky="w")
            ctk.CTkLabel(self.fact_table, text=f["proveedor_nombre"], text_color=COLOR_TEXT, font=ctk.CTkFont(size=11)).grid(row=i, column=1, padx=6, pady=4, sticky="w")
            ctk.CTkLabel(self.fact_table, text=f["descripcion"][:25], text_color="#bdc3c7", font=ctk.CTkFont(size=11)).grid(row=i, column=2, padx=6, pady=4, sticky="w")
            ctk.CTkLabel(self.fact_table, text=moneda, text_color="#8e44ad" if moneda == "USD" else COLOR_TEXT, font=ctk.CTkFont(size=11, weight="bold")).grid(row=i, column=3, padx=6, pady=4, sticky="w")
            ctk.CTkLabel(self.fact_table, text=f["fecha_compra"], text_color="#bdc3c7", font=ctk.CTkFont(size=11)).grid(row=i, column=4, padx=6, pady=4, sticky="w")
            ctk.CTkLabel(self.fact_table, text=f"{simbolo} {f['monto_base']:,.0f}", text_color=COLOR_TEXT, font=ctk.CTkFont(size=11)).grid(row=i, column=5, padx=6, pady=4, sticky="w")
            ctk.CTkLabel(self.fact_table, text=f"{simbolo} {f['iva_monto']:,.0f}", text_color="#bdc3c7", font=ctk.CTkFont(size=11)).grid(row=i, column=6, padx=6, pady=4, sticky="w")
            ctk.CTkLabel(self.fact_table, text=f"{simbolo} {f['total']:,.0f}", text_color=COLOR_ACCENT, font=ctk.CTkFont(size=11, weight="bold")).grid(row=i, column=7, padx=6, pady=4, sticky="w")
            ctk.CTkLabel(self.fact_table, text=f["estado"].capitalize(), text_color=fg, font=ctk.CTkFont(size=11, weight="bold")).grid(row=i, column=8, padx=6, pady=4, sticky="w")

            acc_frame = ctk.CTkFrame(self.fact_table, fg_color="transparent")
            acc_frame.grid(row=i, column=9, padx=6, pady=4, sticky="w")

            if f["estado"] == "pendiente":
                ctk.CTkButton(
                    acc_frame, text="💰 Pagar", width=65, height=26,
                    font=ctk.CTkFont(size=10),
                    fg_color="#27ae60", hover_color="#2ecc71",
                    command=lambda fid=f["id"]: self._pagar_factura(fid)
                ).pack(side="left", padx=(0, 4))

                ctk.CTkButton(
                    acc_frame, text="❌", width=28, height=26,
                    font=ctk.CTkFont(size=10),
                    fg_color="#7f8c8d", hover_color="#e74c3c",
                    command=lambda fid=f["id"]: self._cancelar_factura(fid)
                ).pack(side="left")

    def _actualizar_combo_proveedores(self) -> None:
        proveedores = self.proveedor_controller.obtener_proveedores(activos_solo=True)
        nombres = [p["nombre"] for p in proveedores]
        self.combo_proveedor.configure(values=nombres)
        if nombres:
            self.combo_proveedor.set(nombres[0])
        else:
            self.combo_proveedor.set("")

    def _actualizar_resumen(self) -> None:
        resumen = self.proveedor_controller.obtener_resumen_cuentas_por_pagar()
        self.lbl_cant_pendientes.configure(text=str(resumen["cantidad_pendientes"]))
        self.lbl_total_uyu.configure(text=f"$ {resumen['total_uyu']:,.2f}")
        self.lbl_total_usd.configure(text=f"U$S {resumen['total_usd']:,.2f}")

    # ── ACCIONES PROVEEDORES ─────────────────────────────────────────────────

    def _guardar_proveedor(self) -> None:
        nombre = self._prov_entries[0].get().strip()
        if not nombre:
            self._prov_entries[0].configure(border_color=COLOR_ERROR, border_width=2)
            return
        self._prov_entries[0].configure(border_color=COLOR_SECONDARY, border_width=0)

        telefono = self._prov_entries[1].get().strip()
        email = self._prov_entries[2].get().strip()
        direccion = self._prov_entries[3].get().strip()

        try:
            self.proveedor_controller.crear_proveedor(nombre, telefono, email, direccion)
            for e in self._prov_entries:
                e.delete(0, "end")
            self._render_proveedores()
            self._actualizar_combo_proveedores()
        except Exception as e:
            print(f"Error al guardar proveedor: {e}")

    def _toggle_proveedor(self, proveedor_id: int) -> None:
        self.proveedor_controller.toggle_proveedor_estado(proveedor_id)
        self._render_proveedores()
        self._actualizar_combo_proveedores()

    # ── ACCIONES FACTURAS ────────────────────────────────────────────────────

    def _recalcular(self, event=None) -> None:
        try:
            monto = float(self.entry_fact_monto.get() or "0")
            iva_pct = float(self.entry_fact_iva_pct.get() or "0")
            iva_monto = monto * (iva_pct / 100)
            total = monto + iva_monto
            moneda = self.combo_moneda.get()
            simbolo = "U$S" if moneda == "USD" else "$"
            self.lbl_iva_monto.configure(text=f"{simbolo} {iva_monto:,.2f}")
            self.lbl_fact_total.configure(text=f"{simbolo} {total:,.2f}")
        except ValueError:
            pass

    def _guardar_factura(self) -> None:
        nombre_prov = self.combo_proveedor.get().strip()
        if not nombre_prov:
            self.combo_proveedor.configure(border_color=COLOR_ERROR, border_width=2)
            return
        self.combo_proveedor.configure(border_color=COLOR_SECONDARY, border_width=0)

        proveedores = self.proveedor_controller.obtener_proveedores(activos_solo=True)
        proveedor = next((p for p in proveedores if p["nombre"] == nombre_prov), None)
        if not proveedor:
            return

        descripcion = self.entry_fact_desc.get().strip()
        if not descripcion:
            self.entry_fact_desc.configure(border_color=COLOR_ERROR, border_width=2)
            return
        self.entry_fact_desc.configure(border_color=COLOR_SECONDARY, border_width=0)

        fecha_compra = self.entry_fact_fecha.get().strip()
        if not fecha_compra:
            self.entry_fact_fecha.configure(border_color=COLOR_ERROR, border_width=2)
            return
        self.entry_fact_fecha.configure(border_color=COLOR_SECONDARY, border_width=0)

        try:
            monto_base = float(self.entry_fact_monto.get() or "0")
            iva_pct = float(self.entry_fact_iva_pct.get() or "0")
        except ValueError:
            self.entry_fact_monto.configure(border_color=COLOR_ERROR, border_width=2)
            return
        self.entry_fact_monto.configure(border_color=COLOR_SECONDARY, border_width=0)

        iva_monto = monto_base * (iva_pct / 100)
        total = monto_base + iva_monto
        moneda = self.combo_moneda.get()
        fecha_pago = self.entry_fact_fecha_pago.get().strip()
        obs = self.entry_fact_obs.get().strip()

        try:
            self.proveedor_controller.crear_factura(
                proveedor_id=proveedor["id"],
                descripcion=descripcion,
                fecha_compra=fecha_compra,
                monto_base=monto_base,
                iva_porcentaje=iva_pct,
                iva_monto=iva_monto,
                total=total,
                moneda=moneda,
                fecha_pago=fecha_pago,
                observaciones=obs,
            )
            self.entry_fact_desc.delete(0, "end")
            self.entry_fact_monto.delete(0, "end")
            self.entry_fact_fecha_pago.delete(0, "end")
            self.entry_fact_obs.delete(0, "end")
            self.lbl_iva_monto.configure(text="$ 0.00")
            self.lbl_fact_total.configure(text="$ 0.00")
            self._actualizar_resumen()
            self._filtrar_facturas()
        except Exception as e:
            print(f"Error al guardar factura: {e}")

    def _pagar_factura(self, factura_id: int) -> None:
        dialog = ctk.CTkToplevel(self)
        dialog.title("Confirmar Pago")
        dialog.geometry("350x180")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        self.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() // 2) - 175
        y = self.winfo_rooty() + (self.winfo_height() // 2) - 90
        dialog.geometry(f"+{x}+{y}")

        ctk.CTkLabel(dialog, text="¿Marcar esta factura como pagada?", font=ctk.CTkFont(size=14), text_color=COLOR_TEXT).pack(pady=(25, 10))

        fecha_hoy = datetime.now().strftime("%Y-%m-%d")
        ctk.CTkLabel(dialog, text=f"Fecha de pago: {fecha_hoy}", font=ctk.CTkFont(size=12), text_color="#bdc3c7").pack(pady=(0, 15))

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack()

        def confirmar():
            self.proveedor_controller.marcar_factura_pagada(factura_id, fecha_hoy)
            dialog.destroy()
            self._actualizar_resumen()
            self._filtrar_facturas()

        ctk.CTkButton(btn_frame, text="✅ Sí, pagar", width=120, fg_color="#27ae60", hover_color="#2ecc71", command=confirmar).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="Cancelar", width=100, fg_color="#7f8c8d", hover_color="#95a5a6", command=dialog.destroy).pack(side="left", padx=10)

    def _cancelar_factura(self, factura_id: int) -> None:
        dialog = ctk.CTkToplevel(self)
        dialog.title("Cancelar Factura")
        dialog.geometry("350x150")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        self.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() // 2) - 175
        y = self.winfo_rooty() + (self.winfo_height() // 2) - 75
        dialog.geometry(f"+{x}+{y}")

        ctk.CTkLabel(dialog, text="¿Cancelar esta factura?", font=ctk.CTkFont(size=14), text_color=COLOR_ERROR).pack(pady=(25, 15))

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack()

        def confirmar():
            self.proveedor_controller.cancelar_factura(factura_id)
            dialog.destroy()
            self._actualizar_resumen()
            self._filtrar_facturas()

        ctk.CTkButton(btn_frame, text="❌ Sí, cancelar", width=120, fg_color="#e74c3c", hover_color="#c0392b", command=confirmar).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="Volver", width=100, fg_color="#7f8c8d", hover_color="#95a5a6", command=dialog.destroy).pack(side="left", padx=10)

    def _filtrar_facturas(self, valor=None) -> None:
        filtro = self.combo_filtro.get()
        estado = None if filtro == "Todas" else filtro.lower()
        self._render_facturas(estado=estado)

    # ── EXPORTAR EXCEL ───────────────────────────────────────────────────────

    def _exportar_facturas_excel(self) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            facturas = self.proveedor_controller.obtener_facturas()
            if not facturas:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Facturas Proveedores"

            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            pend_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
            paga_fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
            canc_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            headers = ["ID", "Proveedor", "Descripción", "Moneda", "Fecha Compra", "Monto Base", "IVA %", "IVA Monto", "Total", "Estado", "Fecha Pago", "Observaciones"]
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            total_uyu = 0.0
            total_usd = 0.0
            for i, f in enumerate(facturas, start=2):
                moneda = f.get("moneda", "UYU") or "UYU"
                row_data = [
                    f["id"], f["proveedor_nombre"], f["descripcion"], moneda,
                    f["fecha_compra"], f["monto_base"], f["iva_porcentaje"],
                    f["iva_monto"], f["total"], f["estado"].capitalize(),
                    f["fecha_pago"] or "-", f["observaciones"] or "-"
                ]
                fill_map = {"pendiente": pend_fill, "pagada": paga_fill, "cancelada": canc_fill}
                row_fill = fill_map.get(f["estado"], None)

                for col, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=i, column=col, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    if row_fill:
                        cell.fill = row_fill

                if f["estado"] == "pendiente":
                    if moneda == "USD":
                        total_usd += f["total"]
                    else:
                        total_uyu += f["total"]

            total_row = len(facturas) + 3
            ws.cell(row=total_row, column=8, value="TOTAL PENDIENTE UYU:").font = Font(bold=True, size=12)
            ws.cell(row=total_row, column=9, value=total_uyu).font = Font(bold=True, size=12, color="E74C3C")
            ws.cell(row=total_row, column=9).number_format = '#,##0.00'

            ws.cell(row=total_row + 1, column=8, value="TOTAL PENDIENTE USD:").font = Font(bold=True, size=12)
            ws.cell(row=total_row + 1, column=9, value=total_usd).font = Font(bold=True, size=12, color="27AE60")
            ws.cell(row=total_row + 1, column=9).number_format = '#,##0.00'

            anchos = [6, 20, 30, 8, 14, 14, 8, 14, 14, 12, 14, 25]
            for i, w in enumerate(anchos, start=1):
                col_letter = chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
                ws.column_dimensions[col_letter].width = w

            filename = f"facturas_proveedores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            print(f"✅ Excel exportado: {filename}")

        except ImportError:
            print("❌ openpyxl no está instalado. Ejecutá: pip install openpyxl")
        except Exception as e:
            print(f"❌ Error al exportar: {e}")
