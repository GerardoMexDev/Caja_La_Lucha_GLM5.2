import os
from datetime import datetime
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def generar_reporte_diario_excel(
    movimientos: List[Dict[str, Any]],
    resumen: Dict[str, Any],
    caja_info: Dict[str, Any],
) -> str:
    """Genera un archivo Excel con el reporte de caja y devuelve la ruta."""
    # Guardar directamente en la carpeta de Descargas del usuario
    output_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    os.makedirs(output_dir, exist_ok=True)

    fecha_str = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"Reporte_Caja_{fecha_str}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen del Día"

    # --- ESTILOS ---
    header_font = Font(name='Calibri', bold=True, size=14, color="FFFFFF")
    subheader_font = Font(name='Calibri', bold=True, size=11)
    normal_font = Font(name='Calibri', size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    money_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- CABECERA ---
    ws.merge_cells('A1:G1')
    ws['A1'] = "REPORTE DE CAJA - LA LUCHA"
    ws['A1'].font = Font(name='Calibri', bold=True, size=18, color="2C3E50")
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Fecha Apertura: {caja_info.get('fecha_apertura', 'N/A')}"
    ws['A2'].font = normal_font
    ws['A2'].alignment = Alignment(horizontal='center')

    # --- RESUMEN POR MONEDA ---
    row = 4
    ws.cell(row=row, column=1, value="RESUMEN POR MONEDA").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    for col, h in enumerate(["Moneda", "Ingresos", "Egresos", "Saldo Neto"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = subheader_font
        cell.border = thin_border

    for moneda in ['UYU', 'USD', 'BRL']:
        row += 1
        datos = resumen[moneda]
        ws.cell(row=row, column=1, value=moneda).border = thin_border
        c2 = ws.cell(row=row, column=2, value=datos['ingresos'])
        c2.border = thin_border
        c2.number_format = '#,##0.00'
        
        c3 = ws.cell(row=row, column=3, value=datos['egresos'])
        c3.border = thin_border
        c3.number_format = '#,##0.00'
        
        c4 = ws.cell(row=row, column=4, value=datos['saldo'])
        c4.border = thin_border
        c4.number_format = '#,##0.00'
        c4.fill = money_fill

    # --- EFECTIVO VS BANCO ---
    row += 2
    ws.cell(row=row, column=1, value="EFECTIVO VS BANCO (UYU)").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    ws.cell(row=row, column=1, value="Fondo Inicial:").font = subheader_font
    ws.cell(row=row, column=2, value=resumen['fondo_inicial']).number_format = '#,##0.00'
    
    row += 1
    ws.cell(row=row, column=1, value="Efectivo en Caja:").font = subheader_font
    c_efectivo = ws.cell(row=row, column=2, value=resumen['efectivo_uyu'])
    c_efectivo.number_format = '#,##0.00'
    c_efectivo.fill = money_fill

    row += 1
    ws.cell(row=row, column=1, value="Dinero en Banco:").font = subheader_font
    c_banco = ws.cell(row=row, column=2, value=resumen['banco_uyu'])
    c_banco.number_format = '#,##0.00'
    c_banco.fill = money_fill

    # --- DETALLE DE MOVIMIENTOS ---
    row += 2
    ws.cell(row=row, column=1, value="DETALLE DE MOVIMIENTOS").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(f'A{row}:G{row}')

    row += 1
    mov_headers = ["Hora", "Tipo", "Categoría", "Motivo", "Monto", "Moneda", "Método Pago"]
    for col, h in enumerate(mov_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = subheader_font
        cell.border = thin_border

    for mov in movimientos:
        row += 1
        # Extraer solo la hora si viene con fecha
        hora = mov['created_at'].split(' ')[1] if ' ' in str(mov['created_at']) else str(mov['created_at'])
        
        ws.cell(row=row, column=1, value=hora).border = thin_border
        ws.cell(row=row, column=2, value=mov['tipo'].upper()).border = thin_border
        ws.cell(row=row, column=3, value=mov['categoria']).border = thin_border
        ws.cell(row=row, column=4, value=mov['motivo']).border = thin_border
        
        c_monto = ws.cell(row=row, column=5, value=mov['monto'])
        c_monto.border = thin_border
        c_monto.number_format = '#,##0.00'
        
        ws.cell(row=row, column=6, value=mov['moneda']).border = thin_border
        ws.cell(row=row, column=7, value=mov['metodo_pago']).border = thin_border

    # --- AJUSTAR ANCHOS DE COLUMNA ---
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 20

    wb.save(filepath)
    return filepath